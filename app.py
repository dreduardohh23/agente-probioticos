"""
AGENTE PROBIOTICOS - Web Application
Genomma Lab | Celula de Innovacion en Probioticos
"""

import streamlit as st
import anthropic
import json
import os
import tempfile
import time
from datetime import datetime
from pathlib import Path

# ── Persistent storage for conversations ──
CONV_DIR = Path(os.path.dirname(__file__)) / "saved_conversations"
CONV_DIR.mkdir(exist_ok=True)


def save_conversations_to_disk(conversations):
    """Save all conversations to a JSON file on server."""
    data = {}
    for name, conv in conversations.items():
        data[name] = {
            "fecha": conv.get("fecha", ""),
            "messages": [{"role": m["role"], "content": m.get("display_content", m["content"])} for m in conv["messages"]]
        }
    with open(CONV_DIR / "conversations.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_conversations_from_disk():
    """Load conversations from server JSON file."""
    path = CONV_DIR / "conversations.json"
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            result = {}
            for name, conv in data.items():
                result[name] = {
                    "fecha": conv.get("fecha", ""),
                    "messages": conv["messages"]
                }
            return result
        except Exception:
            return {}
    return {}

# ── Page config ──
st.set_page_config(
    page_title="Agente Probioticos | Genomma Lab",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS - Dark Scientific Theme ──
st.markdown("""
<style>
    /* ── Global Dark Theme ── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    html, body, [data-testid="stAppViewContainer"] {
        background-color: #0A0E1A !important;
        color: #E2E8F0 !important;
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0D1321 0%, #111827 100%) !important;
        border-right: 1px solid rgba(0, 212, 170, 0.15) !important;
        min-width: 310px !important;
    }

    div[data-testid="stSidebarContent"] {
        padding-top: 1rem;
    }

    /* ── Sidebar collapse/expand button — make it VERY visible ── */
    button[data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"],
    button[kind="header"] {
        background-color: #00D4AA !important;
        color: #0A0E1A !important;
        border: 2px solid #00D4AA !important;
        border-radius: 8px !important;
        width: 40px !important;
        height: 40px !important;
        top: 12px !important;
        left: 12px !important;
        z-index: 9999 !important;
        opacity: 1 !important;
        box-shadow: 0 2px 10px rgba(0, 212, 170, 0.4) !important;
    }
    button[data-testid="stSidebarCollapsedControl"]:hover,
    [data-testid="collapsedControl"]:hover {
        background-color: #00FFD0 !important;
        box-shadow: 0 4px 20px rgba(0, 212, 170, 0.6) !important;
    }
    button[data-testid="stSidebarCollapsedControl"] svg,
    [data-testid="collapsedControl"] svg {
        fill: #0A0E1A !important;
        stroke: #0A0E1A !important;
    }

    /* ── Header ── */
    .main-header {
        background: linear-gradient(135deg, #0D1321 0%, #1A1F35 40%, #0F2027 100%);
        border: 1px solid rgba(0, 212, 170, 0.2);
        padding: 1.8rem 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        position: relative;
        overflow: hidden;
    }
    .main-header::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent, #00D4AA, #00B4D8, transparent);
    }
    .main-header h1 {
        color: #FFFFFF;
        margin: 0;
        font-size: 1.9rem;
        font-family: 'Inter', sans-serif;
        font-weight: 700;
        letter-spacing: -0.02em;
    }
    .main-header p {
        color: #00D4AA;
        margin: 0.4rem 0 0 0;
        font-size: 0.9rem;
        font-weight: 500;
        letter-spacing: 0.05em;
        text-transform: uppercase;
    }
    .header-badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.7rem;
        font-weight: 600;
        background: rgba(0, 212, 170, 0.15);
        color: #00D4AA;
        border: 1px solid rgba(0, 212, 170, 0.3);
        margin-top: 0.5rem;
    }

    /* ── DNA Helix decoration ── */
    .dna-decoration {
        position: absolute;
        right: 2rem;
        top: 50%;
        transform: translateY(-50%);
        opacity: 0.08;
        font-size: 4rem;
    }

    /* ── Chat Messages ── */
    [data-testid="stChatMessage"] {
        border-radius: 12px !important;
        border: 1px solid rgba(255,255,255,0.05) !important;
        background: #111827 !important;
        margin-bottom: 0.5rem;
    }

    /* ── Buttons ── */
    .stButton > button {
        background: linear-gradient(135deg, #0D2137 0%, #132D46 100%) !important;
        border: 1px solid rgba(0, 212, 170, 0.25) !important;
        color: #E2E8F0 !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        transition: all 0.3s ease !important;
    }
    .stButton > button:hover {
        border-color: #00D4AA !important;
        box-shadow: 0 0 15px rgba(0, 212, 170, 0.15) !important;
        color: #00D4AA !important;
    }

    /* ── Download buttons ── */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #064E3B 0%, #065F46 100%) !important;
        border: 1px solid rgba(0, 212, 170, 0.3) !important;
        color: #A7F3D0 !important;
        border-radius: 8px !important;
    }

    /* ── Input fields ── */
    .stTextInput input, .stSelectbox select, [data-testid="stTextInput"] input {
        background: #1A1F35 !important;
        border: 1px solid rgba(0, 212, 170, 0.2) !important;
        color: #E2E8F0 !important;
        border-radius: 8px !important;
    }
    .stTextInput input:focus, [data-testid="stTextInput"] input:focus {
        border-color: #00D4AA !important;
        box-shadow: 0 0 0 2px rgba(0, 212, 170, 0.1) !important;
    }

    /* ── Chat Input ── */
    [data-testid="stChatInput"] {
        background: #111827 !important;
        border: 1px solid rgba(0, 212, 170, 0.2) !important;
        border-radius: 12px !important;
    }
    [data-testid="stChatInput"] textarea {
        color: #E2E8F0 !important;
    }

    /* ── Expanders ── */
    [data-testid="stExpander"] {
        background: #111827 !important;
        border: 1px solid rgba(0, 212, 170, 0.1) !important;
        border-radius: 8px !important;
    }

    /* ── Status containers ── */
    [data-testid="stStatusWidget"] {
        background: #111827 !important;
        border: 1px solid rgba(0, 180, 216, 0.2) !important;
    }

    /* ── Metrics / info cards ── */
    .info-card {
        background: linear-gradient(135deg, #111827 0%, #1A1F35 100%);
        border: 1px solid rgba(0, 212, 170, 0.12);
        border-radius: 10px;
        padding: 1.2rem;
        margin-bottom: 0.5rem;
    }
    .info-card h4 {
        color: #00D4AA;
        font-size: 0.85rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin: 0 0 0.5rem 0;
    }
    .info-card p {
        color: #94A3B8;
        font-size: 0.85rem;
        margin: 0.2rem 0;
        line-height: 1.5;
    }

    /* ── Spinner ── */
    .stSpinner > div {
        border-top-color: #00D4AA !important;
    }

    /* ── Success / Error / Warning ── */
    [data-testid="stAlert"] {
        border-radius: 8px !important;
    }

    /* ── Dividers ── */
    hr {
        border-color: rgba(0, 212, 170, 0.1) !important;
    }

    /* ── Sidebar sections ── */
    .sidebar-section-title {
        color: #00D4AA;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin-bottom: 0.5rem;
    }

    /* ── Scrollbar ── */
    ::-webkit-scrollbar { width: 6px; }
    ::-webkit-scrollbar-track { background: #0A0E1A; }
    ::-webkit-scrollbar-thumb { background: #1E293B; border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #00D4AA; }

    /* ── File uploader ── */
    [data-testid="stFileUploader"] {
        background: #111827 !important;
        border: 1px dashed rgba(0, 212, 170, 0.25) !important;
        border-radius: 8px !important;
    }

    /* ── Selectbox ── */
    [data-testid="stSelectbox"] > div > div {
        background: #1A1F35 !important;
        border-color: rgba(0, 212, 170, 0.2) !important;
    }

    /* ── Markdown links ── */
    a { color: #00B4D8 !important; }
    a:hover { color: #00D4AA !important; }

    /* ── Remove Streamlit branding ── */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ── System prompt with all agent knowledge ──
SYSTEM_PROMPT = """Eres el AGENTE PROBIOTICOS de Genomma Lab, una celula de trabajo dedicada al desarrollo e innovacion en la categoria de probioticos.

## Tu Proposito
Definir posicionamiento y portafolio de probioticos diferenciados, con formulas innovadoras y validacion con consumidor, para los mercados de Mexico, LATAM y USA.

## Conocimiento Base del Agente

### PRODUCTO 1: PROBIOTICO + BERBERINA (First-to-Market Global)
- Formula: 6 cepas (~3.3x10^10 UFC) + Berberina HCl 1000mg/dia
- Cepas: B. longum 35624 (1x10^9), B. lactis HN019 (1x10^10), L. acidophilus La-5/NCFM (5x10^9), L. plantarum 299v (1x10^10), L. rhamnosus GG (1x10^10), B. lactis BB-12 (1x10^9)
- Indicacion: SII + DM2 + Resistencia a insulina
- Estudio PREMOTE (Zhang Y et al., Nature Communications, 2020, n=409): BBR+Probiotico redujo HbA1c -1.04% vs -0.71% BBR sola
- 5 niveles de sinergia: farmacocinetico (dhBBR 5x biodisponibilidad), AGCC/GLP-1, barrera intestinal, eje intestino-higado, inmunomodulacion
- NO existe producto combinado comercial relevante a nivel global

### PRODUCTO 2: PROBIOTICO + VITAMINAS (SII + Bienestar)
- Formula: 4 cepas (~1.5x10^10 UFC) + Vitamina D3 2000 UI + Zinc bisglicinato 15mg
- Cepas: L. rhamnosus GG (5x10^9), L. plantarum 299v (5x10^9), B. lactis BB-12 (3x10^9), B. longum 35624 (2x10^9)
- Evidencia: Costanzo 2023 meta-analisis (VitD+probioticos), Akhtar 2022 (Zinc+LGG), Laliani 2023 (VitD+probioticos en SII)
- Alternativa: Psicobioticos + Complejo B (L. helveticus R0052 + B. longum R0175 + B6/B9/B12) para eje intestino-cerebro

### PRODUCTO 3: GOMITA PROBIOTICA
- B. coagulans GanedenBC30 (2x10^9 UFC) - cepa esporulada termoestable
- Formato accesible/lifestyle

### EVIDENCIA CIENTIFICA - CEPAS NIVEL 1
SII: B. longum 35624, L. plantarum 299v, L. rhamnosus GG, VSL#3
Estrenimiento: B. lactis BB-12, B. lactis HN019, B. lactis DN-173 010
Antiinflamatorio: VSL#3 (Nivel 1 AGA), E. coli Nissle 1917, B. longum 35624
Metabolismo glucosa: L. acidophilus La-5 (HbA1c -0.53%), L. rhamnosus GG (HOMA-IR -15%), Akkermansia muciniphila (HOMA-IR -28.6%)

### REGULACION
Mexico (COFEPRIS): Suplemento alimenticio (2-4 meses, $5-15K USD). ALERTA: Berberina en zona gris.
USA (FDA): Dietary supplement DSHEA (3-6 meses, $20-50K USD). Berberina permitida.
Brasil (ANVISA): Mas exigente, 8-18 meses. Colombia (INVIMA): 6-12 meses. Argentina (ANMAT): 6-12 meses.
Estrategia: Fase 1 USA+MX, Fase 2 Peru/Chile/Colombia, Fase 3 Argentina/Brasil.

### COMPETIDORES
USA: Culturelle (LGG, $18-30), Align (B.longum 35624, $25-45), Seed ($49.99 premium), Florastor (S.boulardii)
Mexico: Floratil (lider farmacia), Enterogermina (Sanofi, pediatria), BioGaia, Yakult
GAP: Nadie tiene probiotico+berberina. Mercado MX anclado en "diarrea". Gomitas vacias en farmacia LATAM.

### PROVEEDORES
Cepas: Novonesis (#1, Dinamarca), IFF (#2, USA), Kerry (termoestables), Probi (299v), BioGaia (L.reuteri)
Berberina: Sabinsa (Phytosome, mayor biodisponibilidad), China (>90% produccion global)
CMOs Mexico: Liomont, Medix, Neolpharma. USA: Best Formulations, Catalent. LATAM: Procaps

### FORMAS FARMACEUTICAS
TOP 1: Capsula DRcaps liberacion retardada (supervivencia 80-90%, claim diferenciador)
TOP 2: Sachet monodosis (alta estabilidad, culturalmente aceptado en LATAM)
TOP 3: Gomita con esporas (B. coagulans, mayor crecimiento en USA)
ALERTA: Berberina es antimicrobiana - requiere separacion fisica del probiotico

## Instrucciones de Comportamiento
1. Responde siempre en espanol
2. Basa tus respuestas en evidencia cientifica. Cita estudios cuando sea relevante
3. Cuando busques en la web, prioriza: PubMed, ClinicalTrials.gov, sitios de COFEPRIS/FDA/ANVISA, revistas indexadas
4. Para regulacion, siempre verifica la informacion mas reciente
5. Cuando el usuario suba archivos, analizalos en contexto del proyecto de probioticos
6. Si no estas seguro de algo, indicalo y sugiere donde buscar la informacion
7. Mantente enfocado en el proyecto de Genomma Lab
8. Puedes proponer mejoras, nuevas combinaciones, o alertar sobre riesgos
"""

# ── Helper functions ──

def extract_text_from_pdf(file_bytes):
    """Extract text from PDF file."""
    try:
        from PyPDF2 import PdfReader
        import io
        reader = PdfReader(io.BytesIO(file_bytes))
        text = ""
        for page in reader.pages[:50]:  # Limit to 50 pages
            text += page.extract_text() or ""
        return text[:50000]  # Limit text length
    except Exception as e:
        return f"Error leyendo PDF: {str(e)}"


def extract_text_from_docx(file_bytes):
    """Extract text from DOCX file."""
    try:
        from docx import Document
        import io
        doc = Document(io.BytesIO(file_bytes))
        text = "\n".join([p.text for p in doc.paragraphs])
        return text[:50000]
    except Exception as e:
        return f"Error leyendo DOCX: {str(e)}"


def extract_text_from_xlsx(file_bytes):
    """Extract text from XLSX file."""
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n=== Hoja: {sheet_name} ===\n"
            for row in ws.iter_rows(max_row=200, values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip(" |"):
                    text += row_text + "\n"
        return text[:50000]
    except Exception as e:
        return f"Error leyendo XLSX: {str(e)}"


def extract_file_content(uploaded_file):
    """Extract text content from uploaded file."""
    file_bytes = uploaded_file.read()
    name = uploaded_file.name.lower()

    if name.endswith('.pdf'):
        return extract_text_from_pdf(file_bytes)
    elif name.endswith('.docx'):
        return extract_text_from_docx(file_bytes)
    elif name.endswith(('.xlsx', '.xls')):
        return extract_text_from_xlsx(file_bytes)
    elif name.endswith(('.txt', '.md', '.csv', '.json')):
        return file_bytes.decode('utf-8', errors='ignore')[:50000]
    else:
        return f"Formato no soportado: {name}. Soportados: PDF, DOCX, XLSX, TXT, MD, CSV, JSON"


def search_web(query, max_results=5):
    """Search the web using DuckDuckGo."""
    try:
        from duckduckgo_search import DDGS
        results = []
        with DDGS() as ddgs:
            for r in ddgs.text(query, max_results=max_results):
                results.append({
                    "title": r.get("title", ""),
                    "url": r.get("href", ""),
                    "snippet": r.get("body", "")
                })
        return results
    except Exception as e:
        return [{"title": "Error en busqueda", "url": "", "snippet": str(e)}]


def search_pubmed(query, max_results=5):
    """Search PubMed for scientific articles."""
    try:
        import requests
        base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
        # Search
        search_url = f"{base_url}/esearch.fcgi"
        params = {
            "db": "pubmed",
            "term": query,
            "retmax": max_results,
            "retmode": "json",
            "sort": "relevance"
        }
        resp = requests.get(search_url, params=params, timeout=10)
        data = resp.json()
        ids = data.get("esearchresult", {}).get("idlist", [])

        if not ids:
            return []

        # Fetch details
        fetch_url = f"{base_url}/esummary.fcgi"
        fetch_params = {
            "db": "pubmed",
            "id": ",".join(ids),
            "retmode": "json"
        }
        resp = requests.get(fetch_url, params=fetch_params, timeout=10)
        fetch_data = resp.json()

        results = []
        for uid in ids:
            article = fetch_data.get("result", {}).get(uid, {})
            if article:
                authors = ", ".join([a.get("name", "") for a in article.get("authors", [])[:3]])
                results.append({
                    "title": article.get("title", "Sin titulo"),
                    "authors": authors,
                    "journal": article.get("source", ""),
                    "year": article.get("pubdate", "")[:4],
                    "pmid": uid,
                    "url": f"https://pubmed.ncbi.nlm.nih.gov/{uid}/"
                })
        return results
    except Exception as e:
        return [{"title": f"Error buscando PubMed: {str(e)}", "authors": "", "journal": "", "year": "", "pmid": "", "url": ""}]


def fetch_webpage(url):
    """Fetch and extract text from a webpage."""
    try:
        import requests
        from bs4 import BeautifulSoup
        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
        resp = requests.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(resp.text, 'html.parser')
        # Remove scripts and styles
        for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
            tag.decompose()
        text = soup.get_text(separator='\n', strip=True)
        return text[:30000]
    except Exception as e:
        return f"Error accediendo a {url}: {str(e)}"


def process_tool_call(tool_name, tool_input):
    """Process tool calls from Claude."""
    if tool_name == "web_search":
        results = search_web(tool_input.get("query", ""), tool_input.get("max_results", 5))
        return json.dumps(results, ensure_ascii=False)
    elif tool_name == "pubmed_search":
        results = search_pubmed(tool_input.get("query", ""), tool_input.get("max_results", 5))
        return json.dumps(results, ensure_ascii=False)
    elif tool_name == "fetch_url":
        text = fetch_webpage(tool_input.get("url", ""))
        return text
    return "Herramienta no reconocida"


# ── Tool definitions for Claude ──
TOOLS = [
    {
        "name": "web_search",
        "description": "Busca informacion en la web. Usa para buscar regulaciones, competidores, proveedores, noticias del mercado de probioticos, informacion de COFEPRIS, FDA, ANVISA, etc.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "Termino de busqueda. Ejemplo: 'COFEPRIS regulacion suplementos alimenticios probioticos 2025'"
                },
                "max_results": {
                    "type": "integer",
                    "description": "Numero maximo de resultados (default 5)",
                    "default": 5
                }
            },
            "required": ["query"]
        }
    },
    {
        "name": "pubmed_search",
        "description": "Busca articulos cientificos en PubMed/NCBI. Usa para encontrar estudios clinicos, meta-analisis, revisiones sistematicas sobre probioticos, berberina, vitaminas, etc.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "Termino de busqueda PubMed. Ejemplo: 'berberine probiotics type 2 diabetes randomized controlled trial'"
                },
                "max_results": {
                    "type": "integer",
                    "description": "Numero maximo de resultados (default 5)",
                    "default": 5
                }
            },
            "required": ["query"]
        }
    },
    {
        "name": "fetch_url",
        "description": "Accede a una pagina web y extrae su contenido. Usa para leer paginas de regulaciones, articulos completos, sitios de proveedores, etc.",
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "URL completa de la pagina a leer"
                }
            },
            "required": ["url"]
        }
    }
]


def call_claude_with_retry(client, modelo, system, tools, messages, max_retries=3):
    """Call Claude API with automatic retry on overload/rate limit errors."""
    for attempt in range(max_retries):
        try:
            response = client.messages.create(
                model=modelo,
                max_tokens=4096,
                system=system,
                tools=tools,
                messages=messages
            )
            return response
        except (anthropic.OverloadedError, anthropic.RateLimitError) as e:
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 5  # 5s, 10s, 15s
                st.toast(f"⏳ Servidor ocupado. Reintentando en {wait_time}s... (intento {attempt + 2}/{max_retries})")
                time.sleep(wait_time)
            else:
                raise e
        except anthropic.APIStatusError as e:
            if e.status_code == 529 and attempt < max_retries - 1:
                wait_time = (attempt + 1) * 5
                st.toast(f"⏳ Servidor sobrecargado. Reintentando en {wait_time}s... (intento {attempt + 2}/{max_retries})")
                time.sleep(wait_time)
            else:
                raise e


def get_ai_response(messages, api_key, modelo="claude-3-5-sonnet-20241022"):
    """Get response from Claude with tool use support and auto-retry."""
    client = anthropic.Anthropic(api_key=api_key)

    # Initial call with retry
    response = call_claude_with_retry(client, modelo, SYSTEM_PROMPT, TOOLS, messages)

    # Handle tool use loop
    max_iterations = 5
    iteration = 0

    while response.stop_reason == "tool_use" and iteration < max_iterations:
        iteration += 1
        tool_results = []

        for block in response.content:
            if block.type == "tool_use":
                with st.status(f"🔍 Usando herramienta: **{block.name}**", expanded=True) as status:
                    if block.name == "web_search":
                        st.write(f"Buscando: *{block.input.get('query', '')}*")
                    elif block.name == "pubmed_search":
                        st.write(f"Buscando en PubMed: *{block.input.get('query', '')}*")
                    elif block.name == "fetch_url":
                        st.write(f"Accediendo: *{block.input.get('url', '')}*")

                    result = process_tool_call(block.name, block.input)

                    if block.name in ["web_search", "pubmed_search"]:
                        try:
                            parsed = json.loads(result)
                            for r in parsed[:3]:
                                title = r.get('title', '')
                                st.write(f"  - {title}")
                        except:
                            pass

                    status.update(label=f"✅ {block.name} completado", state="complete")

                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result
                })

        # Continue conversation with tool results
        messages = messages + [
            {"role": "assistant", "content": response.content},
            {"role": "user", "content": tool_results}
        ]

        response = call_claude_with_retry(client, modelo, SYSTEM_PROMPT, TOOLS, messages)

    # Extract final text
    final_text = ""
    for block in response.content:
        if hasattr(block, "text"):
            final_text += block.text

    return final_text


# ═══════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Configuracion")

    # ── Team authentication ──
    TEAM_PASSWORD = st.secrets.get("TEAM_PASSWORD", "")
    API_KEY_SECRET = st.secrets.get("ANTHROPIC_API_KEY", "")

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    st.markdown("---")

    if API_KEY_SECRET and TEAM_PASSWORD:
        # Mode: team access with password
        if not st.session_state.authenticated:
            pwd = st.text_input("Password del equipo", type="password", placeholder="Ingresa el password")
            if pwd:
                if pwd.strip() == TEAM_PASSWORD:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("❌ Password incorrecto")
        if st.session_state.authenticated:
            st.success("✅ Acceso autorizado")
            api_key = API_KEY_SECRET
        else:
            api_key = ""
    else:
        # Fallback: manual API key (for local dev or if secrets not configured)
        api_key_raw = st.text_input(
            "Anthropic API Key",
            type="password",
            placeholder="sk-ant-...",
            help="Obtener en console.anthropic.com"
        )
        api_key = api_key_raw.strip() if api_key_raw else ""
        if api_key:
            st.success("✅ API Key configurada")
        else:
            st.warning("⚠️ Ingresa tu API Key para comenzar")

    # Model selector
    modelo = st.selectbox(
        "Modelo Claude",
        ["claude-3-5-sonnet-20241022", "claude-sonnet-4-20250514", "claude-3-haiku-20240307"],
        index=0,
        help="Si un modelo da error, prueba otro"
    )

    st.markdown("---")

    # Report download
    st.markdown("### 📄 Reporte Estrategico")
    report_path = os.path.join(os.path.dirname(__file__), "Agente_Probioticos_Reporte_Estrategico.docx")
    if os.path.exists(report_path):
        with open(report_path, "rb") as f:
            st.download_button(
                label="⬇️ Descargar Reporte (.docx)",
                data=f.read(),
                file_name="Agente_Probioticos_Reporte_Estrategico.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
        st.caption("Reporte compilado con toda la investigacion")
    else:
        st.info("Reporte no disponible")

    st.markdown("---")

    # File upload
    st.markdown("### 📎 Subir Archivos")
    uploaded_files = st.file_uploader(
        "Arrastra archivos aqui",
        accept_multiple_files=True,
        type=["pdf", "docx", "xlsx", "xls", "txt", "md", "csv", "json"],
        help="Soporta: PDF, DOCX, XLSX, TXT, MD, CSV, JSON"
    )

    if uploaded_files:
        st.markdown(f"**{len(uploaded_files)} archivo(s) cargado(s):**")
        for f in uploaded_files:
            size_kb = f.size / 1024
            st.markdown(f"- 📄 {f.name} ({size_kb:.0f} KB)")

    st.markdown("---")

    # Quick actions
    st.markdown("### 🚀 Acciones Rapidas")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔬 PubMed", use_container_width=True, help="Buscar en PubMed"):
            st.session_state.quick_action = "pubmed"
        if st.button("📋 Regulacion", use_container_width=True, help="Consultar regulacion"):
            st.session_state.quick_action = "regulacion"
    with col2:
        if st.button("🏭 Proveedores", use_container_width=True, help="Buscar proveedores"):
            st.session_state.quick_action = "proveedores"
        if st.button("📊 Competencia", use_container_width=True, help="Analizar competencia"):
            st.session_state.quick_action = "competencia"

    st.markdown("---")

    # Knowledge base summary
    with st.expander("📚 Base de Conocimiento", expanded=False):
        st.markdown("""
        **Productos definidos:**
        - P1: Probiotico + Berberina
        - P2: Probiotico + VitD + Zinc
        - P3: Gomita B. coagulans

        **Cepas clave:** B. longum 35624, L. plantarum 299v, L. rhamnosus GG, B. lactis HN019/BB-12

        **Mercados:** MX, LATAM, USA

        **Competidores:** Floratil, Enterogermina, Culturelle, Align, Seed
        """)

    # ── Conversation management ──
    st.markdown("---")
    st.markdown("### 💾 Conversaciones")

    # Load saved conversations from disk on first run
    if "saved_conversations" not in st.session_state:
        st.session_state.saved_conversations = load_conversations_from_disk()

    # Save current conversation
    if st.session_state.get("messages"):
        save_name = st.text_input("Nombre para guardar", placeholder="Ej: Investigacion berberina")
        if st.button("💾 Guardar conversacion", use_container_width=True):
            if save_name:
                st.session_state.saved_conversations[save_name] = {
                    "messages": list(st.session_state.messages),
                    "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")
                }
                save_conversations_to_disk(st.session_state.saved_conversations)
                st.success(f"✅ Guardada: {save_name}")
            else:
                st.warning("Escribe un nombre para guardar")

    # Load saved conversation
    if st.session_state.saved_conversations:
        conv_names = list(st.session_state.saved_conversations.keys())
        selected = st.selectbox("Conversaciones guardadas", ["-- Seleccionar --"] + conv_names)
        if selected != "-- Seleccionar --":
            fecha = st.session_state.saved_conversations[selected].get("fecha", "")
            if fecha:
                st.caption(f"Guardada: {fecha}")
            col_load, col_del = st.columns(2)
            with col_load:
                if st.button("📂 Cargar", use_container_width=True):
                    st.session_state.messages = list(st.session_state.saved_conversations[selected]["messages"])
                    st.rerun()
            with col_del:
                if st.button("🗑️ Borrar", use_container_width=True):
                    del st.session_state.saved_conversations[selected]
                    save_conversations_to_disk(st.session_state.saved_conversations)
                    st.rerun()

    # Export all conversations as JSON
    if st.session_state.saved_conversations:
        export_data = {}
        for name, conv in st.session_state.saved_conversations.items():
            export_data[name] = {
                "fecha": conv["fecha"],
                "messages": [{"role": m["role"], "content": m.get("display_content", m["content"])} for m in conv["messages"]]
            }
        st.download_button(
            "⬇️ Exportar todas (.json)",
            data=json.dumps(export_data, ensure_ascii=False, indent=2),
            file_name="conversaciones_probioticos.json",
            mime="application/json",
            use_container_width=True
        )

    # Import conversations
    imported = st.file_uploader("Importar conversaciones", type=["json"], key="import_conv")
    if imported:
        try:
            data = json.loads(imported.read().decode("utf-8"))
            for name, conv in data.items():
                st.session_state.saved_conversations[name] = {
                    "messages": conv["messages"],
                    "fecha": conv.get("fecha", "importado")
                }
            save_conversations_to_disk(st.session_state.saved_conversations)
            st.success(f"✅ {len(data)} conversacion(es) importada(s)")
        except Exception as e:
            st.error(f"Error al importar: {e}")

    # New / clear conversation
    st.markdown("---")
    if st.button("🆕 Nueva conversacion", use_container_width=True):
        st.session_state.messages = []
        st.rerun()


# ═══════════════════════════════════════════════════════════
# MAIN AREA
# ═══════════════════════════════════════════════════════════

# Header
st.markdown("""
<div class="main-header">
    <div class="dna-decoration">🧬</div>
    <h1>⚗️ Agente Probioticos</h1>
    <p>Celula de Innovacion Cientifica &bull; Genomma Lab &bull; R&D Probioticos</p>
    <span class="header-badge">● SISTEMA ACTIVO — IA + PubMed + Web Search</span>
</div>
""", unsafe_allow_html=True)

# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []

if "file_contexts" not in st.session_state:
    st.session_state.file_contexts = {}

# Process uploaded files
if uploaded_files:
    for f in uploaded_files:
        if f.name not in st.session_state.file_contexts:
            with st.spinner(f"Procesando {f.name}..."):
                content = extract_file_content(f)
                st.session_state.file_contexts[f.name] = content

# ── TABS: Chat + Dashboard ──
tab_chat, tab_dash = st.tabs(["🧬 Chat Agente", "📊 Dashboard Estrategico"])

with tab_dash:
    st.markdown("""
    <style>
        .kpi-card{background:linear-gradient(145deg,#111827,#0D1B2A);border:1px solid rgba(0,212,170,0.15);border-radius:12px;padding:1.2rem 1.5rem;text-align:center;position:relative;overflow:hidden}
        .kpi-card::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px;background:var(--accent,#00D4AA)}
        .kpi-value{font-size:2rem;font-weight:800;color:var(--accent,#00D4AA);line-height:1}
        .kpi-label{color:#94A3B8;font-size:0.75rem;font-weight:600;text-transform:uppercase;letter-spacing:0.08em;margin-top:0.4rem}
        .kpi-detail{color:#64748B;font-size:0.7rem;margin-top:0.2rem}
        .section-hdr{background:linear-gradient(135deg,#111827,#0D1B2A);border-left:3px solid #00D4AA;border-radius:0 8px 8px 0;padding:0.8rem 1.2rem;margin:1.5rem 0 1rem 0}
        .section-hdr h3{color:#E2E8F0;font-size:1rem;font-weight:700;margin:0}
        .dtable{width:100%;border-collapse:separate;border-spacing:0;background:#111827;border-radius:10px;overflow:hidden;border:1px solid rgba(255,255,255,0.05)}
        .dtable th{background:linear-gradient(135deg,#0D2137,#132D46);color:#00D4AA;font-size:0.7rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;padding:0.8rem 1rem;text-align:left;border-bottom:2px solid rgba(0,212,170,0.3)}
        .dtable td{color:#CBD5E1;font-size:0.8rem;padding:0.7rem 1rem;border-bottom:1px solid rgba(255,255,255,0.03)}
        .dtable tr:hover td{background:rgba(0,212,170,0.05)}
        .dbadge{display:inline-block;padding:0.15rem 0.6rem;border-radius:20px;font-size:0.65rem;font-weight:700}
        .bg{background:rgba(16,185,129,0.15);color:#10B981;border:1px solid rgba(16,185,129,0.3)}
        .bb{background:rgba(59,130,246,0.15);color:#3B82F6;border:1px solid rgba(59,130,246,0.3)}
        .ba{background:rgba(245,158,11,0.15);color:#F59E0B;border:1px solid rgba(245,158,11,0.3)}
        .br{background:rgba(239,68,68,0.15);color:#EF4444;border:1px solid rgba(239,68,68,0.3)}
        .bp{background:rgba(139,92,246,0.15);color:#8B5CF6;border:1px solid rgba(139,92,246,0.3)}
        .ccard{background:#111827;border:1px solid rgba(255,255,255,0.05);border-radius:12px;padding:1.5rem}
        .ctitle{color:#E2E8F0;font-size:0.85rem;font-weight:700;margin-bottom:1rem}
        .bar-row{display:flex;align-items:center;margin-bottom:0.6rem}
        .bar-label{color:#94A3B8;font-size:0.75rem;width:140px;flex-shrink:0}
        .bar-track{flex:1;background:rgba(255,255,255,0.05);border-radius:6px;height:24px;overflow:hidden}
        .bar-fill{height:100%;border-radius:6px;display:flex;align-items:center;padding-left:0.5rem;font-size:0.7rem;font-weight:700;color:white}
        .alert-box{background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.25);border-left:3px solid #EF4444;border-radius:0 8px 8px 0;padding:1rem 1.2rem;margin:0.8rem 0}
        .alert-box .at{color:#EF4444;font-weight:700;font-size:0.8rem}
        .alert-box .ax{color:#CBD5E1;font-size:0.78rem;margin-top:0.3rem}
        .obox{background:rgba(0,212,170,0.05);border:1px solid rgba(0,212,170,0.2);border-radius:10px;padding:1rem 1.2rem;margin-bottom:0.8rem}
        .obox .ot{color:#00D4AA;font-weight:700;font-size:0.85rem}
        .obox .ox{color:#CBD5E1;font-size:0.78rem;margin-top:0.3rem}
    </style>
    """, unsafe_allow_html=True)

    # KPIs
    k1,k2,k3,k4,k5 = st.columns(5)
    with k1:
        st.markdown('<div class="kpi-card" style="--accent:#00D4AA;"><div class="kpi-value">3</div><div class="kpi-label">Productos Pipeline</div><div class="kpi-detail">Prob+BBR | Prob+Vit | Gomita</div></div>', unsafe_allow_html=True)
    with k2:
        st.markdown('<div class="kpi-card" style="--accent:#3B82F6;"><div class="kpi-value">$65-70B</div><div class="kpi-label">Mercado Global</div><div class="kpi-detail">Probioticos 2024 | CAGR 7-9%</div></div>', unsafe_allow_html=True)
    with k3:
        st.markdown('<div class="kpi-card" style="--accent:#8B5CF6;"><div class="kpi-value">6</div><div class="kpi-label">Cepas Nivel 1</div><div class="kpi-detail">Evidencia clinica alta</div></div>', unsafe_allow_html=True)
    with k4:
        st.markdown('<div class="kpi-card" style="--accent:#F59E0B;"><div class="kpi-value">8+</div><div class="kpi-label">Mercados Target</div><div class="kpi-detail">MX, USA, BR, CO, PE, CL, AR</div></div>', unsafe_allow_html=True)
    with k5:
        st.markdown('<div class="kpi-card" style="--accent:#EF4444;"><div class="kpi-value">0</div><div class="kpi-label">Competidores BBR+Prob</div><div class="kpi-detail">First-mover advantage</div></div>', unsafe_allow_html=True)

    # PORTAFOLIO
    st.markdown('<div class="section-hdr"><h3>🧬 PORTAFOLIO PROPUESTO</h3></div>', unsafe_allow_html=True)
    p1,p2,p3 = st.columns(3)
    with p1:
        st.markdown("""<div class="ccard"><div class="ctitle">PRODUCTO 1: Probiotico + Berberina</div><span class="dbadge bg">FIRST-TO-MARKET</span><br><br>
        <table class="dtable"><tr><td style="color:#94A3B8;">Formula</td><td>6 cepas (3.3x10<sup>10</sup> UFC) + BBR 1000mg</td></tr>
        <tr><td style="color:#94A3B8;">Indicacion</td><td>SII + DM2 + Resistencia insulina</td></tr>
        <tr><td style="color:#94A3B8;">Forma</td><td>Sachet monodosis / DRcaps</td></tr>
        <tr><td style="color:#94A3B8;">Precio</td><td><span style="color:#00D4AA;font-weight:700;">$15-25 USD/mes</span></td></tr></table><br>
        <div style="color:#94A3B8;font-size:0.7rem;">EVIDENCIA CLAVE</div>
        <div style="color:#CBD5E1;font-size:0.75rem;margin-top:0.3rem;">PREMOTE (Nature Comms, 2020, n=409): HbA1c <span style="color:#00D4AA;font-weight:700;">-1.04%</span> combinado vs -0.71% BBR sola</div></div>""", unsafe_allow_html=True)
    with p2:
        st.markdown("""<div class="ccard"><div class="ctitle">PRODUCTO 2: Probiotico + Vitaminas</div><span class="dbadge bb">ALTO NIVEL EVIDENCIA</span><br><br>
        <table class="dtable"><tr><td style="color:#94A3B8;">Formula</td><td>4 cepas (1.5x10<sup>10</sup> UFC) + VitD3 + Zinc</td></tr>
        <tr><td style="color:#94A3B8;">Indicacion</td><td>SII + Inmunidad + Bienestar</td></tr>
        <tr><td style="color:#94A3B8;">Forma</td><td>Capsula DRcaps (lib. retardada)</td></tr>
        <tr><td style="color:#94A3B8;">Precio</td><td><span style="color:#3B82F6;font-weight:700;">$18-28 USD/mes</span></td></tr></table><br>
        <div style="color:#94A3B8;font-size:0.7rem;">SINERGIA TRIPLE</div>
        <div style="color:#CBD5E1;font-size:0.75rem;margin-top:0.3rem;">VitD+Prob: PCR <span style="color:#3B82F6;font-weight:700;">-1.28 mg/L</span><br>Zinc+LGG: Restaura barrera intestinal</div></div>""", unsafe_allow_html=True)
    with p3:
        st.markdown("""<div class="ccard"><div class="ctitle">PRODUCTO 3: Gomita Probiotica</div><span class="dbadge bp">LIFESTYLE / ACCESIBLE</span><br><br>
        <table class="dtable"><tr><td style="color:#94A3B8;">Formula</td><td>B. coagulans GBI-30 (2x10<sup>9</sup> UFC)</td></tr>
        <tr><td style="color:#94A3B8;">Indicacion</td><td>Salud digestiva general</td></tr>
        <tr><td style="color:#94A3B8;">Forma</td><td>Gomita (gummy)</td></tr>
        <tr><td style="color:#94A3B8;">Precio</td><td><span style="color:#8B5CF6;font-weight:700;">$12-18 USD/mes</span></td></tr></table><br>
        <div style="color:#94A3B8;font-size:0.7rem;">VENTAJA CLAVE</div>
        <div style="color:#CBD5E1;font-size:0.75rem;margin-top:0.3rem;">Cepa esporulada termoestable<br>Mayor crecimiento en gummy USA<br>Gap en farmacia LATAM</div></div>""", unsafe_allow_html=True)

    # EVIDENCIA CIENTIFICA
    st.markdown('<div class="section-hdr"><h3>🔬 EVIDENCIA CIENTIFICA — CEPAS NIVEL 1</h3></div>', unsafe_allow_html=True)
    st.markdown("""<table class="dtable"><thead><tr><th>Cepa</th><th>Indicacion</th><th>Dosis UFC/dia</th><th>Nivel</th><th>Efecto demostrado</th></tr></thead><tbody>
    <tr><td style="color:#00D4AA;font-weight:600;">B. longum 35624</td><td>SII (todos)</td><td>1x10<sup>9</sup></td><td><span class="dbadge bg">Nivel 1</span></td><td>Reduce IL-6, TNF-alfa, CRP</td></tr>
    <tr><td style="color:#00D4AA;font-weight:600;">L. plantarum 299v</td><td>SII-D, SII-M</td><td>1x10<sup>10</sup></td><td><span class="dbadge bg">Nivel 1</span></td><td>71% vs 11% placebo. FBG -12 mg/dL</td></tr>
    <tr><td style="color:#00D4AA;font-weight:600;">L. rhamnosus GG</td><td>SII-D + Metabolismo</td><td>1-2x10<sup>10</sup></td><td><span class="dbadge bg">Nivel 1</span></td><td>HOMA-IR -15%. Restaura barrera</td></tr>
    <tr><td style="color:#00D4AA;font-weight:600;">B. lactis HN019</td><td>Estrenimiento</td><td>1.8x10<sup>10</sup></td><td><span class="dbadge bg">Nivel 1</span></td><td>Transito colonico -31 horas</td></tr>
    <tr><td style="color:#00D4AA;font-weight:600;">B. lactis BB-12</td><td>Estrenimiento + DM2</td><td>1-10x10<sup>9</sup></td><td><span class="dbadge bg">Nivel 1</span></td><td>HbA1c -0.29%</td></tr>
    <tr><td style="color:#00D4AA;font-weight:600;">L. acidophilus La-5</td><td>DM2</td><td>5x10<sup>9</sup></td><td><span class="dbadge bg">Nivel 1</span></td><td>HbA1c -0.53%. Convierte BBR (5x)</td></tr>
    </tbody></table>""", unsafe_allow_html=True)

    # SINERGIA + PREMOTE
    st.markdown('<div class="section-hdr"><h3>⚡ SINERGIA BBR + PROBIOTICO & ESTUDIO PREMOTE</h3></div>', unsafe_allow_html=True)
    s1,s2 = st.columns(2)
    with s1:
        st.markdown("""<div class="ccard"><div class="ctitle">5 Niveles de Sinergia</div>
        <div class="bar-row"><div class="bar-label">1. Farmacocinetico</div><div class="bar-track"><div class="bar-fill" style="width:95%;background:linear-gradient(90deg,#00D4AA,#00B4D8);">5x biodisponibilidad</div></div></div>
        <div class="bar-row"><div class="bar-label">2. AGCC / GLP-1</div><div class="bar-track"><div class="bar-fill" style="width:85%;background:linear-gradient(90deg,#3B82F6,#8B5CF6);">Estimula insulina</div></div></div>
        <div class="bar-row"><div class="bar-label">3. Barrera intestinal</div><div class="bar-track"><div class="bar-fill" style="width:80%;background:linear-gradient(90deg,#10B981,#059669);">Reduce endotoxemia</div></div></div>
        <div class="bar-row"><div class="bar-label">4. Eje int-higado</div><div class="bar-track"><div class="bar-fill" style="width:75%;background:linear-gradient(90deg,#F59E0B,#EF4444);">Metabolismo hepatico</div></div></div>
        <div class="bar-row"><div class="bar-label">5. Inmunomodulacion</div><div class="bar-track"><div class="bar-fill" style="width:70%;background:linear-gradient(90deg,#8B5CF6,#EC4899);">Reduce inflamacion</div></div></div>
        </div>""", unsafe_allow_html=True)
    with s2:
        st.markdown("""<div class="ccard"><div class="ctitle">Resultados PREMOTE (n=409)</div>
        <table class="dtable"><thead><tr><th>Variable</th><th>BBR sola</th><th>Prob solo</th><th>BBR+Prob</th></tr></thead><tbody>
        <tr><td>HbA1c</td><td>-0.71%</td><td>-0.42%</td><td style="color:#00D4AA;font-weight:700;">-1.04%</td></tr>
        <tr><td>FBG</td><td>-0.99 mmol/L</td><td>-0.58 mmol/L</td><td style="color:#00D4AA;font-weight:700;">-1.49 mmol/L</td></tr>
        <tr><td>Bifidobacterium</td><td><span class="dbadge br">Deplecion</span></td><td><span class="dbadge bb">Aumento</span></td><td><span class="dbadge bg">Restauracion</span></td></tr>
        </tbody></table>
        <div style="color:#64748B;font-size:0.7rem;margin-top:0.8rem;">Zhang Y et al. Nature Communications. 2020;11:5015. RCT, DM2. p&lt;0.05</div></div>""", unsafe_allow_html=True)

    # COMPETENCIA
    st.markdown('<div class="section-hdr"><h3>📊 ANALISIS COMPETITIVO</h3></div>', unsafe_allow_html=True)
    c1,c2 = st.columns(2)
    with c1:
        st.markdown("""<div class="ccard"><div class="ctitle">Competidores USA — Precio USD/mes</div>
        <div class="bar-row"><div class="bar-label">Seed</div><div class="bar-track"><div class="bar-fill" style="width:100%;background:linear-gradient(90deg,#8B5CF6,#EC4899);">$49.99</div></div></div>
        <div class="bar-row"><div class="bar-label">Align</div><div class="bar-track"><div class="bar-fill" style="width:72%;background:linear-gradient(90deg,#3B82F6,#6366F1);">$25-45</div></div></div>
        <div class="bar-row"><div class="bar-label">Culturelle</div><div class="bar-track"><div class="bar-fill" style="width:50%;background:linear-gradient(90deg,#00D4AA,#00B4D8);">$18-30</div></div></div>
        <div class="bar-row"><div class="bar-label">Florastor</div><div class="bar-track"><div class="bar-fill" style="width:48%;background:linear-gradient(90deg,#10B981,#059669);">$20-28</div></div></div>
        <div class="bar-row"><div class="bar-label" style="color:#00D4AA;font-weight:700;">Genomma (P1)</div><div class="bar-track"><div class="bar-fill" style="width:42%;background:linear-gradient(90deg,#F59E0B,#EF4444);">$15-25</div></div></div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown("""<div class="ccard"><div class="ctitle">Competidores Mexico</div>
        <table class="dtable"><thead><tr><th>Marca</th><th>Cepa</th><th>Posicionamiento</th></tr></thead><tbody>
        <tr><td style="font-weight:600;">Floratil</td><td>S. boulardii</td><td>Lider farmacia, diarrea</td></tr>
        <tr><td style="font-weight:600;">Enterogermina</td><td>B. clausii</td><td>Pediatria, Sanofi</td></tr>
        <tr><td style="font-weight:600;">BioGaia</td><td>L. reuteri</td><td>Pediatria premium</td></tr>
        <tr><td style="font-weight:600;">Yakult</td><td>L. casei Shirota</td><td>Masivo, bebida</td></tr>
        </tbody></table><br>
        <div class="obox"><div class="ot">GAP IDENTIFICADO</div><div class="ox">Mercado MX anclado en "probiotico = diarrea". Indicaciones metabolicas VACIAS. Nadie tiene probiotico + berberina.</div></div></div>""", unsafe_allow_html=True)

    # REGULACION
    st.markdown('<div class="section-hdr"><h3>📋 MARCO REGULATORIO</h3></div>', unsafe_allow_html=True)
    st.markdown("""<table class="dtable"><thead><tr><th>Criterio</th><th>Mexico (COFEPRIS)</th><th>USA (FDA)</th><th>Brasil (ANVISA)</th></tr></thead><tbody>
    <tr><td>Clasificacion</td><td>Suplemento alimenticio</td><td>Dietary supplement</td><td>Suplemento alimentar</td></tr>
    <tr><td>Tiempo registro</td><td><span class="dbadge bg">2-4 meses</span></td><td><span class="dbadge bb">3-6 meses</span></td><td><span class="dbadge br">8-18 meses</span></td></tr>
    <tr><td>Costo</td><td>$5-15K USD</td><td>$20-50K USD</td><td>$15-40K USD</td></tr>
    <tr><td>Berberina</td><td><span class="dbadge ba">ZONA GRIS</span></td><td><span class="dbadge bg">PERMITIDA</span></td><td><span class="dbadge ba">CONSULTAR</span></td></tr>
    </tbody></table>""", unsafe_allow_html=True)
    st.markdown("""<div class="alert-box"><div class="at">⚠️ ALERTA: BERBERINA EN MEXICO</div><div class="ax">COFEPRIS puede considerar la berberina como ingrediente farmacologicamente activo. Requiere opinion tecnica previa. En USA no hay problema (historial pre-DSHEA).</div></div>""", unsafe_allow_html=True)

    # PROVEEDORES + TIMELINE
    st.markdown('<div class="section-hdr"><h3>🏭 PROVEEDORES & TIMELINE</h3></div>', unsafe_allow_html=True)
    t1,t2 = st.columns(2)
    with t1:
        st.markdown("""<div class="ccard"><div class="ctitle">Proveedores Principales</div>
        <table class="dtable"><thead><tr><th>Categoria</th><th>Proveedor</th><th>Especialidad</th></tr></thead><tbody>
        <tr><td style="color:#00D4AA;">Cepas</td><td style="font-weight:600;">Novonesis (#1)</td><td>BB-12, HN019, La-5</td></tr>
        <tr><td style="color:#00D4AA;">Cepas</td><td style="font-weight:600;">IFF (#2)</td><td>LGG, NCFM</td></tr>
        <tr><td style="color:#00D4AA;">Cepas</td><td style="font-weight:600;">Probi</td><td>L. plantarum 299v</td></tr>
        <tr><td style="color:#F59E0B;">Berberina</td><td style="font-weight:600;">Sabinsa</td><td>Phytosome (mayor biodis.)</td></tr>
        <tr><td style="color:#8B5CF6;">CMO MX</td><td style="font-weight:600;">Liomont/Medix</td><td>Manufactura local</td></tr>
        <tr><td style="color:#8B5CF6;">CMO USA</td><td style="font-weight:600;">Catalent</td><td>Capsulas, sachets</td></tr>
        </tbody></table></div>""", unsafe_allow_html=True)
    with t2:
        st.markdown("""<div class="ccard"><div class="ctitle">Timeline de Lanzamiento</div><br>
        <div class="obox"><div class="ot">FASE 1 — Meses 1-6</div><div class="ox"><strong>USA + Mexico</strong><br>Dietary supplement (FDA) + Suplemento alimenticio (COFEPRIS)</div></div>
        <div class="obox"><div class="ot">FASE 2 — Meses 4-12</div><div class="ox"><strong>Peru, Chile, Colombia</strong><br>Suplemento dietario (DIGEMID, ISP, INVIMA)</div></div>
        <div class="obox"><div class="ot">FASE 3 — Meses 8-18</div><div class="ox"><strong>Argentina, Brasil</strong><br>ANMAT + ANVISA (mas exigente)</div></div>
        </div>""", unsafe_allow_html=True)

    # OPORTUNIDADES
    st.markdown('<div class="section-hdr"><h3>🎯 OPORTUNIDADES ESTRATEGICAS</h3></div>', unsafe_allow_html=True)
    o1,o2,o3 = st.columns(3)
    with o1:
        st.markdown('<div class="obox"><div class="ot">🥇 OPORTUNIDAD #1</div><div class="ox">Probiotico + Berberina es espacio VACIO global. First-mover advantage con respaldo de Nature Communications.</div></div>', unsafe_allow_html=True)
    with o2:
        st.markdown('<div class="obox"><div class="ot">🥈 OPORTUNIDAD #2</div><div class="ox">Mercado MX/LATAM anclado en "diarrea". Indicaciones metabolicas y bienestar integral VACIAS.</div></div>', unsafe_allow_html=True)
    with o3:
        st.markdown('<div class="obox"><div class="ot">🥉 OPORTUNIDAD #3</div><div class="ox">Gomitas probioticas en farmacia LATAM vacias. B. coagulans termoestable permite gummy sin cadena de frio.</div></div>', unsafe_allow_html=True)

    st.markdown('<div style="text-align:center;padding:1.5rem;border-top:1px solid rgba(255,255,255,0.05);"><div style="color:#64748B;font-size:0.7rem;">GENOMMA LAB | AGENTE PROBIOTICOS | CONFIDENCIAL | Marzo 2026</div></div>', unsafe_allow_html=True)

with tab_chat:

    # Display chat history
    for msg in st.session_state.messages:
        if msg["role"] == "user":
            with st.chat_message("user"):
                st.markdown(msg["display_content"] if "display_content" in msg else msg["content"])
        else:
            with st.chat_message("assistant", avatar="🧬"):
                st.markdown(msg["content"])

    # Handle quick actions
    quick_prompts = {
        "pubmed": "Busca en PubMed los estudios mas recientes (2024-2026) sobre la combinacion de probioticos con berberina para diabetes tipo 2. Dame un resumen de los hallazgos mas relevantes.",
        "regulacion": "Busca la regulacion mas actualizada de COFEPRIS para suplementos alimenticios con probioticos en Mexico. Verifica si hay actualizaciones recientes sobre berberina.",
        "proveedores": "Busca proveedores actualizados de cepas probioticas (Novonesis, IFF, Kerry) y sus ultimas innovaciones en cepas para salud metabolica.",
        "competencia": "Busca los lanzamientos mas recientes de productos probioticos en Mexico y USA. Analiza las tendencias actuales del mercado."
    }

    if "quick_action" in st.session_state:
        action = st.session_state.quick_action
        if action in quick_prompts:
            prompt = quick_prompts[action]
            del st.session_state.quick_action
            st.session_state.pending_prompt = prompt
            st.rerun()

# Chat input (outside tabs - Streamlit requires chat_input at root level)
if prompt := st.chat_input("Pregunta al Agente Probioticos...") or st.session_state.get("pending_prompt"):
    if "pending_prompt" in st.session_state:
        prompt = st.session_state.pending_prompt
        del st.session_state.pending_prompt

    if not api_key:
        st.error("⚠️ Por favor ingresa tu Anthropic API Key en la barra lateral para comenzar.")
        st.stop()

    # Build user message with file context if any
    full_content = prompt
    display_content = prompt

    # Add file contexts if files are uploaded and this is the first message or files are new
    if st.session_state.file_contexts:
        file_context = "\n\n--- ARCHIVOS ADJUNTOS ---\n"
        for fname, content in st.session_state.file_contexts.items():
            file_context += f"\n### Archivo: {fname}\n{content[:15000]}\n"
        full_content = prompt + file_context
        display_content = prompt + f"\n\n📎 *{len(st.session_state.file_contexts)} archivo(s) adjunto(s)*"

    # Display user message
    with st.chat_message("user"):
        st.markdown(display_content)

    # Add to history
    st.session_state.messages.append({
        "role": "user",
        "content": full_content,
        "display_content": display_content
    })

    # Build messages for API (only send text content)
    api_messages = []
    for msg in st.session_state.messages:
        if isinstance(msg["content"], str):
            api_messages.append({"role": msg["role"], "content": msg["content"]})

    # Get AI response
    with st.chat_message("assistant", avatar="🧬"):
        with st.spinner("Pensando..."):
            try:
                response = get_ai_response(api_messages, api_key, modelo)
                st.markdown(response)

                # Add to history
                st.session_state.messages.append({
                    "role": "assistant",
                    "content": response
                })
            except anthropic.AuthenticationError as e:
                st.error(f"❌ API Key invalida. Verifica tu clave en console.anthropic.com\n\nDetalle: {str(e)}")
            except anthropic.PermissionDeniedError as e:
                st.error(f"❌ Permiso denegado. Tu API Key puede no tener acceso al modelo. Detalle: {str(e)}")
            except anthropic.NotFoundError as e:
                st.error(f"❌ Modelo no encontrado. Detalle: {str(e)}")
            except (anthropic.RateLimitError, anthropic.OverloadedError):
                st.error("⏳ Servidor sobrecargado despues de 3 reintentos. Espera 1 minuto e intenta de nuevo.")
            except anthropic.APIStatusError as e:
                if e.status_code == 529:
                    st.error("⏳ Servidor sobrecargado despues de 3 reintentos. Espera 1 minuto e intenta de nuevo.")
                else:
                    st.error(f"❌ Error API: {str(e)}")
            except Exception as e:
                st.error(f"❌ Error: {type(e).__name__}: {str(e)}")


# ── Welcome message if no chat history ──
if not st.session_state.messages:
    st.markdown("---")

    st.markdown("""
    <div style="text-align: center; padding: 1rem 0;">
        <h3 style="color: #E2E8F0; font-weight: 600;">Sistema de Inteligencia para Innovacion en Probioticos</h3>
        <p style="color: #64748B; font-size: 0.9rem;">Acceso a evidencia cientifica, regulacion, cadena de suministro y analisis de mercado en tiempo real</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown("""
        <div class="info-card">
            <h4>🔬 Evidencia Cientifica</h4>
            <p>• Busqueda en PubMed</p>
            <p>• Estudios clinicos RCT</p>
            <p>• Meta-analisis</p>
            <p>• Validacion de cepas</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="info-card">
            <h4>📋 Regulacion Sanitaria</h4>
            <p>• COFEPRIS (Mexico)</p>
            <p>• FDA (USA)</p>
            <p>• ANVISA / INVIMA</p>
            <p>• Claims permitidos</p>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class="info-card">
            <h4>🏭 Supply Chain</h4>
            <p>• Proveedores de cepas</p>
            <p>• Fabricantes CMO</p>
            <p>• Berberina / Vitaminas</p>
            <p>• Formas farmaceuticas</p>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown("""
        <div class="info-card">
            <h4>📊 Inteligencia de Mercado</h4>
            <p>• Analisis competitivo</p>
            <p>• Precios y tendencias</p>
            <p>• Gaps de mercado</p>
            <p>• Oportunidades</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("""
    <div style="background: #111827; border: 1px solid rgba(0,212,170,0.1); border-radius: 10px; padding: 1.2rem;">
        <p style="color: #00D4AA; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.8rem;">Consultas sugeridas</p>
        <p style="color: #94A3B8; font-size: 0.85rem; margin: 0.4rem 0;">→ Busca los ultimos estudios sobre L. plantarum 299v y SII</p>
        <p style="color: #94A3B8; font-size: 0.85rem; margin: 0.4rem 0;">→ Regulacion COFEPRIS para berberina en suplementos</p>
        <p style="color: #94A3B8; font-size: 0.85rem; margin: 0.4rem 0;">→ Compara precios de probioticos en farmacias de Mexico</p>
        <p style="color: #94A3B8; font-size: 0.85rem; margin: 0.4rem 0;">→ Proveedores de B. coagulans GanedenBC30 para Mexico</p>
        <p style="color: #94A3B8; font-size: 0.85rem; margin: 0.4rem 0;">→ Formas farmaceuticas innovadoras en marcas lideres</p>
    </div>
    """, unsafe_allow_html=True)
